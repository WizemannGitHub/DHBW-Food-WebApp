"""Erstellt einen Gantt-Ablaufplan als Excel-Datei mit tagesgenauen Balken.

Terminstruktur (konsistent mit ablaufplan.typ, KW 20–31):
  Start:           12.05.2026  (KW20, Di)  ← M0
  M1 Anforderungen 20.05.2026  (KW21, Mi)
  M2 Prototyp:     02.06.2026  (KW23, Di)
  M3 PM-Abgabe:    24.06.2026  (KW26, Mi)
  M4 Präsentation: 08.07.2026  (KW28, Mi)
  M5 Webanwendung: 15.07.2026  (KW29, Mi)
  M6 Abnahme:      28.07.2026  (KW31, Di)
  M7 Tech. Abgabe: 03.08.2026  (KW32, Mo) ← Projektende
"""

from datetime import date, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Projektdaten ─────────────────────────────────────────────────────────────
PROJECT_TITLE  = "DHBW Food Web App"
PROJECT_LEADER = "Erik Wizemann"
PROJECT_CLIENT = "Mikka Jenne · Dr. Arno Mielke"

# ── Farben (hex, kein #) ─────────────────────────────────────────────────────
C_HEADER = "1A3A5C"
C_PM     = "2E74B5"
C_ANA    = "C00000"
C_IMP    = "7030A0"
C_DES    = "375623"
C_TEST   = "833C00"
C_KRIT   = "C00000"   # kritischer Pfad → rot
C_NORMAL = "70AD47"   # nicht kritisch  → grün
C_WE_HDR = "AAAAAA"   # Wochenende Header
C_WE_BAR = "E8E8E8"   # Wochenende Zellen in Task-Zeilen

# ── KW-Anker (jeweils Montag der Woche) ─────────────────────────────────────
KW = {
    20: date(2026, 5, 11),
    21: date(2026, 5, 18),
    22: date(2026, 5, 25),
    23: date(2026, 6,  1),
    24: date(2026, 6,  8),
    25: date(2026, 6, 15),
    26: date(2026, 6, 22),
    27: date(2026, 6, 29),
    28: date(2026, 7,  6),
    29: date(2026, 7, 13),
    30: date(2026, 7, 20),
    31: date(2026, 7, 27),
}

def kw_mon(n: int) -> date: return KW[n]
def kw_fri(n: int) -> date: return KW[n] + timedelta(4)

PROJECT_START  = date(2026, 5, 12)   # Di KW20
PROJECT_END    = date(2026, 8,  3)   # Mo KW32 — M7 Tech. Abgabe
ABGABE_PM      = date(2026, 6, 24)   # Mi KW26 — M3
PRAESENTATION  = date(2026, 7,  8)   # Mi KW28 — M4
TECH_ABGABE    = date(2026, 8,  3)   # Mo KW32 — M7

# ── Aufgaben-Termine (konsistent mit ablaufplan.typ KW-Spalten) ──────────────
# 1 · Projektmanagement
s11, e11 = PROJECT_START,  kw_fri(20)
s12, e12 = PROJECT_START,  PROJECT_END
s13, e13 = kw_mon(31),     PROJECT_END

# 2 · Anforderungsanalyse
s21, e21 = PROJECT_START,  kw_fri(21)
s22, e22 = kw_mon(21),     kw_fri(21)
s23, e23 = kw_mon(21),     kw_fri(21)

# 3 · Design
s31, e31 = kw_mon(21),     kw_fri(23)
s32, e32 = kw_mon(21),     kw_fri(22)

# 4 · Implementierung
s41, e41 = kw_mon(23),     kw_fri(26)
s42, e42 = kw_mon(24),     kw_fri(27)
s43, e43 = kw_mon(28),     kw_fri(28)

# 5 · Testing & Deployment
s51, e51 = kw_mon(29),     kw_fri(30)
s52, e52 = kw_mon(30),     PROJECT_END
s53, e53 = kw_mon(30),     PROJECT_END

# ── Meilensteine ─────────────────────────────────────────────────────────────
MILESTONES = [
    (date(2026, 5, 12), "M0 Start"),
    (date(2026, 5, 24), "M1 Anforderungen"),
    (date(2026, 6,  7), "M2 Prototyp"),
    (date(2026, 6, 24), "M3 PM-Abgabe"),
    (date(2026, 7,  8), "M4 Präsentation"),
    (date(2026, 7, 15), "M5 Webanwendung"),
    (date(2026, 7, 28), "M6 Abnahme"),
    (date(2026, 8,  3), "M7 Tech. Abgabe"),
]

# ── Aufgabenliste (Reihenfolge: 1 PM → 2 ANA → 3 DES → 4 IMP → 5 TEST) ─────
# (id, name, start, end, farbe_hex, kritisch, ist_phasen_header)
TASKS = [
    ("",    "1 · Projektmanagement",       None, None, C_PM,   False, True ),
    ("1.1", "Projektplanung",              s11,  e11,  C_PM,   False, False),
    ("1.2", "Projektsteuerung",            s12,  e12,  C_PM,   False, False),
    ("1.3", "Projektabschluss",            s13,  e13,  C_PM,   True,  False),

    ("",    "2 · Anforderungsanalyse",     None, None, C_ANA,  False, True ),
    ("2.1", "Ist-Analyse",                 s21,  e21,  C_ANA,  True,  False),
    ("2.2", "Anforderungen erheben",       s22,  e22,  C_ANA,  True,  False),
    ("2.3", "Risikoanalyse",               s23,  e23,  C_ANA,  False, False),

    ("",    "3 · Design",                  None, None, C_DES,  False, True ),
    ("3.1", "UI/UX Design",                s31,  e31,  C_DES,  False, False),
    ("3.2", "Systemarchitektur",           s32,  e32,  C_DES,  True,  False),

    ("",    "4 · Implementierung",         None, None, C_IMP,  False, True ),
    ("4.1", "Frontend",                    s41,  e41,  C_IMP,  True,  False),
    ("4.2", "Backend",                     s42,  e42,  C_IMP,  True,  False),
    ("4.3", "Datenbank",                   s43,  e43,  C_IMP,  True,  False),

    ("",    "5 · Testing & Deployment",    None, None, C_TEST, False, True ),
    ("5.1", "Testing & QA",                s51,  e51,  C_TEST, True,  False),
    ("5.2", "Deployment",                  s52,  e52,  C_TEST, True,  False),
    ("5.3", "Abgabe & Präsentation",       s53,  e53,  C_TEST, True,  False),
]

PRED_MAP = {
    "1.1": "–",        "1.2": "1.1",      "1.3": "5.3",
    "2.1": "1.1",      "2.2": "2.1",      "2.3": "2.1",
    "3.1": "2.2",      "3.2": "2.1",
    "4.1": "3.1/3.2",  "4.2": "3.2",      "4.3": "4.1/4.2",
    "5.1": "4.3",      "5.2": "5.1",      "5.3": "5.1",
}

# ── Alle Tage inkl. Wochenende im Projektzeitraum ────────────────────────────
all_days: list[date] = []
d = PROJECT_START
while d <= PROJECT_END:
    all_days.append(d)
    d += timedelta(1)

# ── Workbook aufbauen ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt"

COL_ID    = 1
COL_NAME  = 2
COL_DUR   = 3
COL_PRED  = 4
COL_FIRST = 5

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 7
ws.column_dimensions["D"].width = 10
for i, day in enumerate(all_days):
    ws.column_dimensions[get_column_letter(COL_FIRST + i)].width = 3.0

last_col = COL_FIRST + len(all_days) - 1

# ── Style-Helfer ──────────────────────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def set_range_border(ws, r1, r2, c1, c2, color="BBBBBB"):
    """Setzt Außenkanten einer Merge-Range korrekt auf allen Rand-Zellen."""
    thin = Side(style="thin", color=color)
    no   = Side(style=None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = Border(
                left   = thin if c == c1 else no,
                right  = thin if c == c2 else no,
                top    = thin if r == r1 else no,
                bottom = thin if r == r2 else no,
            )

C_COL_HDR = "3B6DAD"  # mittleres Blau für Spaltenbeschriftungen

def align_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def align_left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

def hfont(size=9) -> Font:
    return Font(bold=True, color="FFFFFF", size=size)

def nfont(bold=False, size=9) -> Font:
    return Font(bold=bold, size=size)

# ── Info-Header (Zeilen 1–2): Titel + Projekt-Infos ─────────────────────────
ROW_INFO1 = 1
ROW_INFO2 = 2
ROW_MONTH = 3
ROW_DATE  = 4
ROW_DOW   = 5
ROW_TASKS = 6

# Zeile 1: Projekttitel – volle Breite
ws.row_dimensions[ROW_INFO1].height = 28
ws.merge_cells(start_row=ROW_INFO1, end_row=ROW_INFO1, start_column=COL_ID, end_column=last_col)
c = ws.cell(row=ROW_INFO1, column=COL_ID, value=f"  {PROJECT_TITLE}")
c.fill = fill(C_HEADER)
c.font = Font(bold=True, color="FFFFFF", size=15)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = thin_border()
set_range_border(ws, ROW_INFO1, ROW_INFO1, COL_ID, last_col)

# Zeile 2: drei gleichbreite Segmente (Projektleiter | Auftraggeber | Zeitraum)
ws.row_dimensions[ROW_INFO2].height = 18
seg_size = (last_col - COL_ID + 1) // 3
seg_starts = [COL_ID, COL_ID + seg_size, COL_ID + 2 * seg_size]
seg_ends   = [seg_starts[1] - 1, seg_starts[2] - 1, last_col]
seg_labels = [
    (f" Projektleiter", PROJECT_LEADER),
    (f" Auftraggeber",  PROJECT_CLIENT),
    (f" Zeitraum",      f"{PROJECT_START:%d.%m.%Y} – {PROJECT_END:%d.%m.%Y}"),
]
C_INFO = "254C82"  # etwas helleres Blau für Zeile 2
for (label, value), c1, c2 in zip(seg_labels, seg_starts, seg_ends):
    ws.merge_cells(start_row=ROW_INFO2, end_row=ROW_INFO2, start_column=c1, end_column=c2)
    c = ws.cell(row=ROW_INFO2, column=c1, value=f"{label}:  {value}")
    c.fill = fill(C_INFO)
    c.font = Font(color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    set_range_border(ws, ROW_INFO2, ROW_INFO2, c1, c2)

# ── Spaltenköpfe: Info-Spalten über 3 Kopfzeilen mergen ─────────────────────
for c in range(COL_ID, COL_FIRST):
    ws.merge_cells(start_row=ROW_MONTH, end_row=ROW_DOW, start_column=c, end_column=c)

for col_idx, lbl in {COL_ID: "ID", COL_NAME: "Vorgang / Arbeitspaket",
                     COL_DUR: "Dauer", COL_PRED: "Vorgänger"}.items():
    cell = ws.cell(row=ROW_MONTH, column=col_idx, value=lbl)
    cell.fill = fill(C_COL_HDR)
    cell.font = Font(bold=True, color="FFFFFF", size=8)
    cell.alignment = align_center()
    set_range_border(ws, ROW_MONTH, ROW_DOW, col_idx, col_idx)

# ── Monats-Gruppen ────────────────────────────────────────────────────────────
cur_m, m_start = None, None
month_groups = []
for i, day in enumerate(all_days):
    lbl = day.strftime("%B %Y")
    if lbl != cur_m:
        if cur_m:
            month_groups.append((cur_m, m_start, COL_FIRST + i - 1))
        cur_m, m_start = lbl, COL_FIRST + i
month_groups.append((cur_m, m_start, last_col))

for lbl, c1, c2 in month_groups:
    ws.merge_cells(start_row=ROW_MONTH, end_row=ROW_MONTH, start_column=c1, end_column=c2)
    cell = ws.cell(row=ROW_MONTH, column=c1, value=lbl)
    cell.fill = fill(C_HEADER)
    cell.font = hfont()
    cell.alignment = align_center()
    set_range_border(ws, ROW_MONTH, ROW_MONTH, c1, c2)

# ── Datum- und Wochentag-Zeilen ───────────────────────────────────────────────
ms_set = {ms_date for ms_date, _ in MILESTONES}
WDAY = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

for i, day in enumerate(all_days):
    col = COL_FIRST + i
    is_ms = day in ms_set
    is_we = day.weekday() >= 5

    if is_ms:
        hdr_fill, hdr_color = "B8CCE4", "C00000"
        dow_fill, dow_color = "D6E4F0", "C00000"
    elif is_we:
        hdr_fill, hdr_color = C_WE_HDR, "FFFFFF"
        dow_fill, dow_color = "BBBBBB", "FFFFFF"
    else:
        hdr_fill, hdr_color = C_HEADER, "FFFFFF"
        dow_fill, dow_color = "2C4E80", "FFFFFF"

    cd = ws.cell(row=ROW_DATE, column=col, value=day.strftime("%d.%m"))
    cd.fill = fill(hdr_fill)
    cd.font = Font(bold=True, color=hdr_color, size=7)
    cd.alignment = align_center()
    cd.border = thin_border()

    cw = ws.cell(row=ROW_DOW, column=col, value=WDAY[day.weekday()])
    cw.fill = fill(dow_fill)
    cw.font = Font(color=dow_color, size=7)
    cw.alignment = align_center()
    cw.border = thin_border()

ws.row_dimensions[ROW_MONTH].height = 14
ws.row_dimensions[ROW_DATE].height  = 13
ws.row_dimensions[ROW_DOW].height   = 11

# ── Aufgaben-Zeilen ───────────────────────────────────────────────────────────
def lighten(hex_color: str, amount: int = 110) -> str:
    r = min(255, int(hex_color[0:2], 16) + amount)
    g = min(255, int(hex_color[2:4], 16) + amount)
    b = min(255, int(hex_color[4:6], 16) + amount)
    return f"{r:02X}{g:02X}{b:02X}"

row = ROW_TASKS
for (tid, tname, tstart, tend, tcolor, tkrit, tphase) in TASKS:
    ws.row_dimensions[row].height = 16

    if tphase:
        ws.merge_cells(start_row=row, end_row=row,
                       start_column=COL_ID, end_column=last_col)
        cell = ws.cell(row=row, column=COL_ID, value=tname)
        cell.fill = fill(lighten(tcolor))
        cell.font = Font(bold=True, color=tcolor, size=9)
        cell.alignment = align_left()
        cell.border = thin_border()
        row += 1
        continue

    # Dauer in Arbeitstagen (Mo–Fr)
    dur = sum(1 for d in all_days if tstart and tstart <= d <= tend and d.weekday() < 5)

    for col_idx, val, align_fn, bold in [
        (COL_ID,   tid,                          align_center, True ),
        (COL_NAME, tname,                        align_left,   False),
        (COL_DUR,  f"{dur} AT" if dur else "",   align_center, False),
        (COL_PRED, PRED_MAP.get(tid, "–"),       align_center, False),
    ]:
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = nfont(bold=bold)
        c.alignment = align_fn()
        c.border = thin_border()

    bar_color = C_KRIT if tkrit else C_NORMAL
    for i, day in enumerate(all_days):
        col = COL_FIRST + i
        c = ws.cell(row=row, column=col, value="")
        c.border = thin_border()
        is_we = day.weekday() >= 5
        if tstart and tstart <= day <= tend and not is_we:
            c.fill = fill(bar_color)
        elif is_we:
            c.fill = fill(C_WE_BAR)

    row += 1

# ── Meilenstein-Zeile ─────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 20

for col_idx in [COL_ID, COL_DUR, COL_PRED]:
    c = ws.cell(row=row, column=col_idx, value="")
    c.fill = fill("F0F0F0")
    c.border = thin_border()

c = ws.cell(row=row, column=COL_NAME, value="Meilensteine ◆")
c.fill = fill("F0F0F0")
c.font = Font(bold=True, size=9)
c.alignment = align_left()
c.border = thin_border()

ms_by_date = {ms_date: ms_label for ms_date, ms_label in MILESTONES}
for i, day in enumerate(all_days):
    col = COL_FIRST + i
    c = ws.cell(row=row, column=col, value="")
    c.border = thin_border()
    if day.weekday() >= 5:
        c.fill = fill(C_WE_BAR)
    elif day in ms_by_date:
        c.value = f"◆ {ms_by_date[day]}"
        c.fill = fill("FFF2CC")
        c.font = Font(bold=True, color="C00000", size=7)
        c.alignment = align_center()

# ── Footer ────────────────────────────────────────────────────────────────────
row += 2
ws.merge_cells(start_row=row, end_row=row, start_column=COL_ID, end_column=last_col)
c = ws.cell(row=row, column=COL_ID,
    value="Kritischer Pfad: 2.1 Ist-Analyse → 2.2 Anforderungen → 3.2 Systemarchitektur "
          "→ 4.1 Frontend → 4.3 Datenbank → 5.1 Testing & QA → 5.3 Abgabe & Präsentation → 1.3 Projektabschluss")
c.font = Font(size=8, italic=True)
c.alignment = align_left()

row += 1
ws.merge_cells(start_row=row, end_row=row, start_column=COL_ID, end_column=last_col)
c = ws.cell(row=row, column=COL_ID,
    value="Legende:  ■ Rot = kritischer Pfad   ■ Grün = nicht kritisch   "
          f"◆ Gold = Meilenstein   PM-Abgabe: {ABGABE_PM:%d.%m.%Y}   "
          f"Präsentation: {PRAESENTATION:%d.%m.%Y}   Tech. Abgabe: {TECH_ABGABE:%d.%m.%Y}")
c.font = Font(size=8)
c.alignment = align_left()

# ── Einfrieren: Info-Spalten + Kopfzeilen ────────────────────────────────────
ws.freeze_panes = f"{get_column_letter(COL_FIRST)}{ROW_TASKS}"

# ── Speichern ─────────────────────────────────────────────────────────────────
out = "/Users/I766778/ProjektWebPro/DHBW-Food-WebApp/Dokumente/ablaufplan.xlsx"
wb.save(out)
print(f"Gespeichert: {out}")
print(f"Zeitraum: {PROJECT_START:%d.%m.%Y} – {PROJECT_END:%d.%m.%Y}  ({len(all_days)} Kalendertage)")
print(f"PM-Abgabe: {ABGABE_PM:%d.%m.%Y}  |  Präsentation: {PRAESENTATION:%d.%m.%Y}  |  Tech. Abgabe: {TECH_ABGABE:%d.%m.%Y}")
